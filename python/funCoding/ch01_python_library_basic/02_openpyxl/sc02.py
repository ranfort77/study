
from openpyxl import Workbook, styles, load_workbook


# 엑셀 파일 생성
wb = Workbook()    # 워크북 생성
ws = wb.active     # 워크시트 얻기
ws['A1'] = 'Hello' # A1셀에 Hello 입력
wb.save('snapshot-1.xlsx')

# 새로운 시트 생성
ws2 = wb.create_sheet('새로운 시트', 0)
ws2['A1'] = 100
wb.save('snapshot-2.xlsx')

# 시트 삭제
wb.remove(ws)
wb.save('snapshot-3.xlsx')

# 시트 이름 변경
ws2.title = '이름변경 시트'
wb.save('snapshot-4.xlsx')

# 셀 병합
ws2.merge_cells('A1:E1')
wb.save('snapshot-5.xlsx')

# 셀 폰트, 정렬, 채우기 
c = ws2['A1']
c.font = styles.Font('굴림', 16, True)
c.alignment = styles.Alignment('center', 'center')
c.fill = styles.PatternFill(patternType='solid', 
                            fgColor=styles.Color('FFC000'))
wb.save('snapshot-6.xlsx')

# 데이터 넣기
mat = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
for r, row in enumerate(mat):
    for c, ele in enumerate(row):
        ws2.cell(r + 2, c + 1).value = ele
wb.save('snapshot-7.xlsx')

# 데이터 읽기
wb2 = load_workbook('snapshot-7.xlsx')
ws2 = wb2.active
for row in ws2.rows:
    print(row)
    for ele in row:
        print(ele.value)

wb.close()
wb2.close()