'''
train.xlsx 파일에 있는 데이터를 읽고
new_train.xlsx 에 저장하는 예
'''

import openpyxl

wb = openpyxl.load_workbook('data/train.xlsx')
ws = wb.active

newws = wb.create_sheet('new sheet')

for row in ws.rows:
    for e in row:
        print(type(e.row), type(e.column))
        newws.cell(row=e.row, column=e.col_idx).value = e.value
    
wb.save("data/new_train.xlsx")
wb.close()
